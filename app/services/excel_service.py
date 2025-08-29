from __future__ import annotations
from io import BytesIO
from typing import List, Tuple

from openpyxl import load_workbook

# Expected Excel headers (case-insensitive):
#  Valvula, Cantidad, Ubicación, # de serie, Ficha tecnica, Simbolo
REQUIRED_COLUMNS = ["valvula", "cantidad", "ubicacion", "# de serie", "ficha tecnica", "simbolo"]


def parse_excel_to_rows(data: bytes) -> List[Tuple]:
    """Parse Excel bytes with openpyxl and return tuples for new DB schema.

    Expected columns (case-insensitive):
      Valvula, Cantidad, Ubicación, # de serie, Ficha tecnica, Simbolo

    Returns list of tuples for bulk_insert_valves:
      (id, valvula, cantidad, ubicacion, numero_serie, ficha_tecnica, simbolo)
      id can be None to auto-assign.
    """
    bio = BytesIO(data)
    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    # normalize headers: lowercase and remove accents/minor variants
    def norm_header(h: str) -> str:
        s = str(h or "").strip().lower()
        replacements = {
            "á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n",
        }
        for a,b in replacements.items():
            s = s.replace(a,b)
        return s

    raw_headers = [str(c) if c is not None else "" for c in header_cells]
    headers = [norm_header(c) for c in raw_headers]

    # Map required columns to indices
    idx_map = {}
    for col in REQUIRED_COLUMNS:
        try:
            idx_map[col] = headers.index(norm_header(col))
        except ValueError:
            missing = [c for c in REQUIRED_COLUMNS if norm_header(c) not in headers]
            raise ValueError(f"Excel missing required columns: {', '.join(missing)}")

    def norm(val) -> str:
        if val is None:
            return ""
        return str(val).strip()

    def to_int(x: str) -> int:
        try:
            return int(float(x)) if x else 0
        except Exception:
            raise ValueError(f"Invalid id value: {x}")

    rows: List[Tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if row is None or all(cell is None for cell in row):
            continue
        valvula = norm(row[idx_map["valvula"]]) if idx_map["valvula"] < len(row) else ""
        cantidad_raw = norm(row[idx_map["cantidad"]]) if idx_map["cantidad"] < len(row) else "0"
        ubicacion = norm(row[idx_map["ubicacion"]]) if idx_map["ubicacion"] < len(row) else ""
        numero_serie = norm(row[idx_map["# de serie"]]) if idx_map["# de serie"] < len(row) else ""
        ficha_tecnica = norm(row[idx_map["ficha tecnica"]]) if idx_map["ficha tecnica"] < len(row) else ""
        simbolo = norm(row[idx_map["simbolo"]]) if idx_map["simbolo"] < len(row) else ""

        # id: let DB auto-assign
        id_val = None
        try:
            cantidad = int(float(cantidad_raw)) if cantidad_raw else 0
        except Exception:
            cantidad = 0

        rows.append((id_val, valvula, cantidad, ubicacion, numero_serie, ficha_tecnica, simbolo))

    return rows
